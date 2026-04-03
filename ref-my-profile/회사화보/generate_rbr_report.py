from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from math import isfinite
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


BASE_DIR = Path(r"C:\WorkSpace\report-skills")
SOURCE_XLSX = BASE_DIR / "RBR_ASW26_Datas" / "ASW26 Database (Hyosung TNS).xlsx"
OUTPUT_XLSX = BASE_DIR / f"RBR_지역별실적_{datetime.now():%Y%m}.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="17365D")
SUB_FILL = PatternFill("solid", fgColor="D9E2F3")
LIGHT_FILL = PatternFill("solid", fgColor="F4F7FB")
WHITE_FONT = Font(name="맑은 고딕", color="FFFFFF", bold=True, size=10)
BASE_FONT = Font(name="맑은 고딕", size=10)
TITLE_FONT = Font(name="맑은 고딕", size=13, bold=True)
THIN = Side(style="thin", color="B8C4D0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_source():
    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)

    market = {}
    ws = wb["Market Size and Forecasts"]
    for row in ws.iter_rows(min_row=9, values_only=True):
        year, region, country = row[0], row[1], row[2]
        if region == "North America" and country in {"USA", "Canada", "Total"} and year in {2022, 2025, 2028}:
            market[(year, country)] = {
                "app_multivendor": row[6],
                "app_native": row[7],
                "app_total": row[8],
                "mon_multivendor": row[12],
                "mon_other": row[13],
                "mon_total": row[14],
            }

    vendor_data = defaultdict(dict)
    vendor_headers = {}
    for sheet in ["Application Vendors", "Monitoring Vendors", "Platform Vendors"]:
        ws = wb[sheet]
        headers = list(next(ws.iter_rows(min_row=6, max_row=6, values_only=True)))
        vendor_headers[sheet] = headers
        for row in ws.iter_rows(min_row=7, values_only=True):
            year, measure, region, country = row[0], row[1], row[2], row[3]
            if (
                region == "North America"
                and country in {"USA", "Canada", "Total"}
                and year in {2022, 2025}
                and measure in {"Total", "Multivendor"}
            ):
                vendor_data[sheet][(year, measure, country)] = row
    return market, vendor_data, vendor_headers


def pct_change(current, base):
    if not isinstance(current, (int, float)) or not isinstance(base, (int, float)) or base == 0:
        return None
    return current / base - 1


def safe_ratio(numerator, denominator):
    if not isinstance(numerator, (int, float)) or not isinstance(denominator, (int, float)) or denominator == 0:
        return None
    return numerator / denominator


def top_vendors(row, headers, count=3):
    total = row[4]
    pairs = []
    for idx, value in enumerate(row[5:], start=5):
        if isinstance(value, (int, float)) and value:
            pairs.append((headers[idx], value, value / total))
    pairs.sort(key=lambda x: x[1], reverse=True)
    return pairs[:count]


def set_header(ws, row, values):
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX


def set_subheader(ws, row, text, cols=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = SUB_FILL
    cell.font = Font(name="맑은 고딕", bold=True, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BOX


def style_range(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.font = BASE_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BOX


def autofit(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_idx = col_cells[0].column
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 42)


def set_page(ws, print_area, fit_height=1):
    ws.print_area = print_area
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = fit_height
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2)


def build_workbook():
    market, vendor_data, vendor_headers = load_source()
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("임원용 요약 대시보드")
    ws2 = wb.create_sheet("지역별 상세 실적")
    ws3 = wb.create_sheet("이슈 및 개선과제")
    ws4 = wb.create_sheet("원본 데이터")

    build_sheet1(ws1, market)
    build_sheet2(ws2, market)
    build_sheet3(ws3, market, vendor_data, vendor_headers)
    build_sheet4(ws4, market, vendor_data, vendor_headers)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        autofit(ws)
    set_page(ws1, "A1:J28", fit_height=1)
    set_page(ws2, "A1:J20", fit_height=1)
    set_page(ws3, "A1:H28", fit_height=2)
    set_page(ws4, "A1:Q40", fit_height=2)
    wb.save(OUTPUT_XLSX)


def build_sheet1(ws, market):
    ws["A1"] = "결론: 북미 ATM SW 시장은 2022~2025년 Application 멀티벤더 11.1%, Monitoring 멀티벤더 14.4% 감소로 운영 단순화 기조가 수치로 확인됨"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A1:J1")

    ws["A2"] = "작성일"
    ws["B2"] = "=TODAY()"
    ws["B2"].number_format = "yyyy-mm-dd"
    ws["D2"] = "지표 정의"
    ws["E2"] = "목표=2028 전망치 / 실적=2025 설치기반 / 달성률=실적÷목표 / 증감=2022→2025"
    for ref in ["A2", "D2"]:
        ws[ref].fill = SUB_FILL
        ws[ref].font = Font(name="맑은 고딕", bold=True)
        ws[ref].border = BOX
    for ref in ["B2", "E2"]:
        ws[ref].border = BOX
        ws[ref].font = BASE_FONT

    set_subheader(ws, 4, "핵심 수치", cols=6)
    set_header(ws, 5, ["구분", "2022", "2025", "2028 전망", "증감(22→25)", "증감(25→28 전망)"])
    kpi_rows = [
        ("Application 전체시장", market[(2022, "Total")]["app_total"], market[(2025, "Total")]["app_total"], market[(2028, "Total")]["app_total"]),
        ("Application 멀티벤더", market[(2022, "Total")]["app_multivendor"], market[(2025, "Total")]["app_multivendor"], market[(2028, "Total")]["app_multivendor"]),
        ("Monitoring 전체시장", market[(2022, "Total")]["mon_total"], market[(2025, "Total")]["mon_total"], market[(2028, "Total")]["mon_total"]),
        ("Monitoring 멀티벤더", market[(2022, "Total")]["mon_multivendor"], market[(2025, "Total")]["mon_multivendor"], market[(2028, "Total")]["mon_multivendor"]),
    ]
    row = 6
    for name, v22, v25, v28 in kpi_rows:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=v22)
        ws.cell(row=row, column=3, value=v25)
        ws.cell(row=row, column=4, value=v28)
        ws.cell(row=row, column=5, value=pct_change(v25, v22))
        ws.cell(row=row, column=6, value=pct_change(v28, v25))
        row += 1
    style_range(ws, 5, 9, 1, 6)
    for r in range(6, 10):
        for c in [2, 3, 4]:
            ws.cell(r, c).number_format = "#,##0"
        for c in [5, 6]:
            ws.cell(r, c).number_format = "0.0%"

    set_subheader(ws, 11, "지역별 달성률(2025 실적 ÷ 2028 전망)", cols=5)
    set_header(ws, 12, ["지역", "시장", "목표(2028)", "실적(2025)", "달성률"])
    region_rows = [
        ("USA", "Application", market[(2028, "USA")]["app_total"], market[(2025, "USA")]["app_total"]),
        ("Canada", "Application", market[(2028, "Canada")]["app_total"], market[(2025, "Canada")]["app_total"]),
        ("North America", "Application", market[(2028, "Total")]["app_total"], market[(2025, "Total")]["app_total"]),
        ("USA", "Monitoring", market[(2028, "USA")]["mon_total"], market[(2025, "USA")]["mon_total"]),
        ("Canada", "Monitoring", market[(2028, "Canada")]["mon_total"], market[(2025, "Canada")]["mon_total"]),
        ("North America", "Monitoring", market[(2028, "Total")]["mon_total"], market[(2025, "Total")]["mon_total"]),
    ]
    row = 13
    for region, market_name, target, actual in region_rows:
        ws.cell(row=row, column=1, value=region)
        ws.cell(row=row, column=2, value=market_name)
        ws.cell(row=row, column=3, value=target)
        ws.cell(row=row, column=4, value=actual)
        ws.cell(row=row, column=5, value=safe_ratio(actual, target))
        row += 1
    style_range(ws, 12, 18, 1, 5)
    for r in range(13, 19):
        ws.cell(r, 3).number_format = "#,##0"
        ws.cell(r, 4).number_format = "#,##0"
        ws.cell(r, 5).number_format = "0.0%"

    set_subheader(ws, 20, "임원 메모", cols=10)
    memos = [
        "미국 Application 멀티벤더는 124,728대에서 113,406대로 감소해 단일벤더화 압력이 확인됨.",
        "캐나다 Application 멀티벤더는 4,313대에서 1,300대로 급감해 우선 관리 지역으로 분류됨.",
        "Monitoring 총시장 점유는 NCR Atleos가 30.7%에서 43.6%로 확대돼 경쟁집중도가 상승함.",
        "BoA·Truist는 Hyosung TNS 하드웨어 34% 기반이 확인돼 소프트웨어 번들 제안 여지가 존재함.",
    ]
    for idx, memo in enumerate(memos, start=21):
        ws.cell(idx, 1, value=f"• {memo}")
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=10)
        ws.cell(idx, 1).border = BOX
        ws.cell(idx, 1).alignment = Alignment(wrap_text=True)
        ws.cell(idx, 1).font = BASE_FONT

    chart = BarChart()
    chart.title = "지역별 달성률"
    chart.y_axis.title = "달성률"
    chart.height = 6
    chart.width = 10
    data = Reference(ws, min_col=5, min_row=12, max_row=18)
    cats = Reference(ws, min_col=1, min_row=13, max_row=18)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.style = 10
    chart.y_axis.numFmt = "0%"
    ws.add_chart(chart, "G4")

    line = LineChart()
    line.title = "기준연도 대비 증감률(2022→2025)"
    line.y_axis.title = "증감률"
    line.height = 6
    line.width = 10
    ws["G25"] = "지역"
    ws["H25"] = "Application"
    ws["I25"] = "Monitoring"
    change_rows = [
        ("USA", pct_change(market[(2025, "USA")]["app_total"], market[(2022, "USA")]["app_total"]), pct_change(market[(2025, "USA")]["mon_total"], market[(2022, "USA")]["mon_total"])),
        ("Canada", pct_change(market[(2025, "Canada")]["app_total"], market[(2022, "Canada")]["app_total"]), pct_change(market[(2025, "Canada")]["mon_total"], market[(2022, "Canada")]["mon_total"])),
        ("North America", pct_change(market[(2025, "Total")]["app_total"], market[(2022, "Total")]["app_total"]), pct_change(market[(2025, "Total")]["mon_total"], market[(2022, "Total")]["mon_total"])),
    ]
    start = 26
    for idx, (region, app_chg, mon_chg) in enumerate(change_rows, start=start):
        ws.cell(idx, 7, value=region)
        ws.cell(idx, 8, value=app_chg)
        ws.cell(idx, 9, value=mon_chg)
        ws.cell(idx, 8).number_format = "0.0%"
        ws.cell(idx, 9).number_format = "0.0%"
    data = Reference(ws, min_col=8, max_col=9, min_row=25, max_row=28)
    cats = Reference(ws, min_col=7, min_row=26, max_row=28)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.style = 12
    line.y_axis.numFmt = "0%"
    ws.add_chart(line, "G11")


def build_sheet2(ws, market):
    set_header(ws, 1, ["지역", "시장", "지표", "목표(2028)", "실적(2025)", "달성률", "증감(22→25)", "2025 점유율", "등급", "근거"])
    rows = []
    for country in ["USA", "Canada", "Total"]:
        label = "North America" if country == "Total" else country
        rows.extend(
            [
                (
                    label,
                    "Application",
                    "전체시장",
                    market[(2028, country)]["app_total"],
                    market[(2025, country)]["app_total"],
                    safe_ratio(market[(2025, country)]["app_total"], market[(2028, country)]["app_total"]),
                    pct_change(market[(2025, country)]["app_total"], market[(2022, country)]["app_total"]),
                    safe_ratio(market[(2025, country)]["app_multivendor"], market[(2025, country)]["app_total"]),
                    "N/A",
                    "ASW26 Database / Market Size and Forecasts",
                ),
                (
                    label,
                    "Application",
                    "멀티벤더",
                    market[(2028, country)]["app_multivendor"],
                    market[(2025, country)]["app_multivendor"],
                    safe_ratio(market[(2025, country)]["app_multivendor"], market[(2028, country)]["app_multivendor"]),
                    pct_change(market[(2025, country)]["app_multivendor"], market[(2022, country)]["app_multivendor"]),
                    safe_ratio(market[(2025, country)]["app_multivendor"], market[(2025, country)]["app_total"]),
                    "N/A",
                    "ASW26 Database / Market Size and Forecasts",
                ),
                (
                    label,
                    "Monitoring",
                    "전체시장",
                    market[(2028, country)]["mon_total"],
                    market[(2025, country)]["mon_total"],
                    safe_ratio(market[(2025, country)]["mon_total"], market[(2028, country)]["mon_total"]),
                    pct_change(market[(2025, country)]["mon_total"], market[(2022, country)]["mon_total"]),
                    safe_ratio(market[(2025, country)]["mon_multivendor"], market[(2025, country)]["mon_total"]),
                    "N/A",
                    "ASW26 Database / Market Size and Forecasts",
                ),
                (
                    label,
                    "Monitoring",
                    "멀티벤더",
                    market[(2028, country)]["mon_multivendor"],
                    market[(2025, country)]["mon_multivendor"],
                    safe_ratio(market[(2025, country)]["mon_multivendor"], market[(2028, country)]["mon_multivendor"]),
                    pct_change(market[(2025, country)]["mon_multivendor"], market[(2022, country)]["mon_multivendor"]),
                    safe_ratio(market[(2025, country)]["mon_multivendor"], market[(2025, country)]["mon_total"]),
                    "N/A",
                    "ASW26 Database / Market Size and Forecasts",
                ),
            ]
        )
    for r_idx, data in enumerate(rows, start=2):
        for c_idx, value in enumerate(data, start=1):
            ws.cell(r_idx, c_idx, value=value)
    style_range(ws, 1, len(rows) + 1, 1, 10)
    for r in range(2, len(rows) + 2):
        for c in [4, 5]:
            ws.cell(r, c).number_format = "#,##0"
        for c in [6, 7, 8]:
            ws.cell(r, c).number_format = "0.0%"
    ws["A16"] = "주석"
    ws["B16"] = "RBR 원천에는 월간 목표/실적/등급 기준이 없어, 목표는 2028 전망치로 대체하고 등급은 미기재함."
    ws["A16"].fill = SUB_FILL
    ws["A16"].font = Font(name="맑은 고딕", bold=True)
    ws["A16"].border = BOX
    ws["B16"].border = BOX
    ws.merge_cells("B16:J16")


def build_sheet3(ws, market, vendor_data, vendor_headers):
    set_subheader(ws, 1, "주요 이슈", cols=8)
    set_header(ws, 2, ["우선순위", "지역", "시장", "이슈", "근거 수치", "개선과제", "기대효과", "출처"])
    issues = [
        (
            1,
            "Canada",
            "Application",
            "멀티벤더 기반 급감",
            "4,313 → 1,300 (-69.9%), 2025 비중 3.7%",
            "단일벤더 전환 계정 식별 및 mixed hardware 레퍼런스 보강",
            "멀티벤더 유지 계정 방어",
            "ASW26 DB / Market Size and Forecasts",
        ),
        (
            2,
            "North America",
            "Monitoring",
            "NCR Atleos 중심 재편",
            "총시장 점유 30.7% → 43.6%",
            "NCR 대항 포지션으로 운영 효율·통합관리 가치 제안 정교화",
            "모니터링 교체 수요 대응",
            "ASW26 DB / Monitoring Vendors",
        ),
        (
            3,
            "USA",
            "Application",
            "멀티벤더 감소에도 대형은행 레퍼런스 유지",
            "BoA HW 34%, Truist HW 34%",
            "하드웨어 설치 기반을 소프트웨어 번들 제안의 진입점으로 활용",
            "대형은행 교차판매 기반 확보",
            "Part 2 정리본",
        ),
        (
            4,
            "USA",
            "Monitoring",
            "Fiserv 비중 하락",
            "멀티벤더 점유 47.8% → 21.2%",
            "교체·전환 계정 추적 및 운영관리 관점 메시지 강화",
            "전환 수요 포착",
            "ASW26 DB / Monitoring Vendors",
        ),
    ]
    for r_idx, data in enumerate(issues, start=3):
        for c_idx, value in enumerate(data, start=1):
            ws.cell(r_idx, c_idx, value=value)
    style_range(ws, 2, 6, 1, 8)
    for r in range(3, 7):
        ws.cell(r, 5).alignment = Alignment(wrap_text=True, vertical="top")

    set_subheader(ws, 8, "임원 보고 문장 톤 분석", cols=8)
    set_header(ws, 9, ["항목", "분석 결과", "실제 적용 문장 예시", "", "", "", "", ""])
    tone_rows = [
        (
            "결론 우선",
            "첫 줄에 시장 방향과 수치를 함께 제시하고, 다음 문장에서 원인을 압축 설명함.",
            "북미 ATM SW 시장은 운영 단순화 기조로 멀티벤더 설치기반이 축소되고 있다.",
        ),
        (
            "문제 정의",
            "시장 둔화 자체보다 운영 구조 변화, 벤더 집중, 멀티벤더 축소를 문제로 정의함.",
            "캐나다는 멀티벤더 기반 급감으로 전략 방어가 필요한 지역이다.",
        ),
        (
            "운영/매출/관리 표현",
            "설치 기반, 점유율, 운영 효율, 네트워크 관리, 번들 제안, 교체 수요 같은 표현을 반복 사용함.",
            "대형 플릿에서 소프트웨어 통합·모니터링 제안 여지가 존재한다.",
        ),
        (
            "개선과제 제시",
            "추상적 비전보다 계정 방어, 레퍼런스 확보, 번들 확대처럼 실행 단위를 직접 적시함.",
            "하드웨어 설치 기반을 소프트웨어 제안의 진입점으로 활용해야 한다.",
        ),
        (
            "기대효과 표현",
            "정량은 대수·점유율·증감률로, 정성은 방어·확대·전환 대응으로 표현함.",
            "멀티벤더 유지 계정 방어와 교체 수요 대응 기반 확보가 가능하다.",
        ),
    ]
    row = 10
    for item, result, example in tone_rows:
        ws.cell(row, 1, value=item)
        ws.cell(row, 2, value=result)
        ws.cell(row, 3, value=example)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        row += 1
    style_range(ws, 9, 14, 1, 8)

    set_subheader(ws, 16, "임원 코멘트용 한 줄 요약", cols=8)
    comments = [
        "북미는 시장 전체보다 멀티벤더 기반의 감소폭이 더 커 소프트웨어 표준화가 빠르게 진행되고 있습니다.",
        "미국은 대형은행 레퍼런스가 유지되지만, 캐나다는 멀티벤더 축소 속도가 빨라 방어 우선순위가 높습니다.",
        "Monitoring은 NCR Atleos 집중도가 높아져 경쟁 포지션 재정의가 필요합니다.",
    ]
    for idx, text in enumerate(comments, start=17):
        ws.cell(idx, 1, value=f"{idx-16}.")
        ws.cell(idx, 2, value=text)
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=8)
        for col in [1, 2]:
            ws.cell(idx, col).border = BOX
        ws.cell(idx, 2).alignment = Alignment(wrap_text=True, vertical="top")


def build_sheet4(ws, market, vendor_data, vendor_headers):
    set_subheader(ws, 1, "시장 규모 원본", cols=12)
    set_header(ws, 2, ["연도", "지역", "Application 멀티벤더", "Application Native", "Application 합계", "Monitoring 멀티벤더", "Monitoring Other", "Monitoring 합계", "출처", "", "", ""])
    row = 3
    for country in ["USA", "Canada", "Total"]:
        label = "North America" if country == "Total" else country
        for year in [2022, 2025, 2028]:
            m = market[(year, country)]
            values = [year, label, m["app_multivendor"], m["app_native"], m["app_total"], m["mon_multivendor"], m["mon_other"], m["mon_total"], "ASW26 DB"]
            for col, val in enumerate(values, start=1):
                ws.cell(row, col, value=val)
            row += 1
    style_range(ws, 2, row - 1, 1, 9)
    for r in range(3, row):
        for c in range(3, 8):
            if isinstance(ws.cell(r, c).value, (int, float)):
                ws.cell(r, c).number_format = "#,##0"

    row += 1
    set_subheader(ws, row, "벤더 점유 원본", cols=17)
    row += 1
    set_header(ws, row, ["시트", "연도", "구분", "지역", "총설치", "1위", "1위 비중", "2위", "2위 비중", "3위", "3위 비중", "출처", "", "", "", "", ""])
    row += 1
    for sheet in ["Application Vendors", "Monitoring Vendors", "Platform Vendors"]:
        headers = vendor_headers[sheet]
        for key, data in vendor_data[sheet].items():
            year, measure, country = key
            if country != "Total":
                continue
            tops = top_vendors(data, headers, count=3)
            values = [
                sheet,
                year,
                measure,
                "North America",
                data[4],
                tops[0][0],
                tops[0][2],
                tops[1][0],
                tops[1][2],
                tops[2][0],
                tops[2][2],
                "ASW26 DB",
            ]
            for col, val in enumerate(values, start=1):
                ws.cell(row, col, value=val)
            row += 1
    style_range(ws, row - 12, row - 1, 1, 12)
    for r in range(row - 12, row):
        ws.cell(r, 5).number_format = "#,##0"
        for c in [7, 9, 11]:
            ws.cell(r, c).number_format = "0.0%"

    row += 1
    set_subheader(ws, row, "고객 현황 원본 요약", cols=8)
    row += 1
    set_header(ws, row, ["고객", "하드웨어", "소프트웨어", "운영체제", "핵심 시사점", "출처", "", ""])
    row += 1
    customers = [
        ("Bank of America", "Hyosung TNS 34%", "멀티벤더", "Windows 10", "대형 플릿 통합·모니터링 제안 여지", "Part 2 정리본"),
        ("Truist", "Hyosung TNS 34%, NCR Atleos 66%", "KAL / NCR Atleos / Hyosung TNS", "Windows 10", "번들 제안 및 협업 포지션 확보 가능", "Part 2 정리본"),
        ("Wells Fargo", "Hyosung TNS 1%", "n/a", "Windows 10", "파일럿 확대형 접근 필요", "Part 2 정리본"),
    ]
    for data in customers:
        for col, val in enumerate(data, start=1):
            ws.cell(row, col, value=val)
        row += 1
    style_range(ws, row - 4, row - 1, 1, 6)


if __name__ == "__main__":
    build_workbook()
    print(OUTPUT_XLSX)
